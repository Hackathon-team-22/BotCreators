from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..domain.reporting import ExcelReport


class ExcelRendererAdapter:
    """Адаптер для рендеринга отчетов в формате Excel.
    """

    _HEADER_FONT = Font(bold=True)
    _HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFEFEFEF")
    _HEADER_ALIGNMENT = Alignment(horizontal="center")
    _THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def render(self, report: ExcelReport) -> bytes:
        wb = Workbook()

        if report.sheets:
            wb.remove(wb.active)

        for sheet in report.sheets:
            ws = wb.create_sheet(title=sheet.name)
            if sheet.columns:
                # Заголовки
                ws.append(list(sheet.columns))
                for cell in ws[1]:
                    cell.font = self._HEADER_FONT
                    cell.fill = self._HEADER_FILL
                    cell.alignment = self._HEADER_ALIGNMENT
                # Заморозка первой строки
                ws.freeze_panes = "A2"

            # Данные
            for row in sheet.rows:
                ws.append([row.get(col, "") for col in sheet.columns])

            if sheet.columns:
                # Включаем автофильтр по всей таблице
                if ws.max_row >= 1 and ws.max_column >= 1:
                    ws.auto_filter.ref = ws.dimensions

                # Тонкие границы для всех заполненных ячеек
                for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row_cells:
                        cell.border = self._THIN_BORDER

                # Подбор ширины столбцов по максимальной длине содержимого и заголовка.
                # Учитываем также кнопку автофильтра — небольшой дополнительный запас.
                for idx, column_name in enumerate(sheet.columns, start=1):
                    column_letter = get_column_letter(idx)
                    max_length = len(str(column_name)) if column_name is not None else 0
                    # Пропускаем первую строку (заголовок уже учтён), смотрим только данные.
                    for cell in ws[column_letter][1:]:
                        if cell.value is None:
                            continue
                        max_length = max(max_length, len(str(cell.value)))
                    # Минимальная ширина, запас под фильтр и ограничение сверху.
                    adjusted = max_length + 4  # +4 — запас под фильтр и отступы
                    adjusted = max(adjusted, 12)
                    ws.column_dimensions[column_letter].width = min(adjusted, 50)

        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()
